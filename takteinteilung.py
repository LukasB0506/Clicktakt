# -*- coding: utf-8 -*-
"""
Takteinteilungs-Tool – PyQt6 Edition
Porsche Leipzig | Dark Mode
"""

import sys, os, json, random, math
from collections import defaultdict
from datetime import date

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QListWidget, QListWidgetItem,
    QTabWidget, QFrame, QSplitter, QScrollArea, QGridLayout,
    QSpinBox, QComboBox, QCheckBox, QMessageBox, QFileDialog,
    QDialog, QDialogButtonBox, QTableWidget, QTableWidgetItem,
    QHeaderView, QRadioButton, QButtonGroup, QGroupBox, QSizePolicy,
    QAbstractItemView
)
from PyQt6.QtCore import Qt, QSize, pyqtSignal
from PyQt6.QtGui import (
    QColor, QFont, QPalette, QIcon, QPainter, QPixmap,
    QLinearGradient, QPen, QBrush, QFontMetrics
)

# ── Imports für Excel ──────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font as XFont, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Pfade ──────────────────────────────────────────────────────
if getattr(sys, "frozen", False):
    _BASIS = os.path.dirname(sys.executable)
else:
    _BASIS = os.path.dirname(os.path.abspath(__file__))

SAVE_FILE = os.path.join(_BASIS, "team_config.json")
CONFIG_VERSION = 3  # Erhöhen wenn sich das Dateiformat ändert

WOCHENTAGE = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]

# ── Porsche Farbpalette ────────────────────────────────────────
P = {
    "bg":        "#111111",
    "bg2":       "#1C1C1C",
    "bg3":       "#252525",
    "bg4":       "#2E2E2E",
    "border":    "#3A3A3A",
    "red":       "#D5001C",
    "red2":      "#FF1A33",
    "white":     "#FFFFFF",
    "text":      "#F0F0F0",
    "text2":     "#AAAAAA",
    "green":     "#27AE60",
    "orange":    "#E67E22",
    "gold":      "#C9A84C",
}

TAKT_PRESET = [
    "#1A3A5C","#1A4A2E","#4A3500","#4A1A1A","#2E1A4A",
    "#1A3A38","#4A2E0E","#3A1A3A","#2A1A40","#1A2E40",
]

QSS = f"""
QMainWindow, QWidget {{
    background-color: {P['bg']};
    color: {P['text']};
    font-family: 'Segoe UI', Arial;
    font-size: 10pt;
}}
QTabWidget::pane {{
    border: none;
    background: {P['bg2']};
}}
QTabWidget::tab-bar {{ alignment: left; }}
QTabBar::tab {{
    background: {P['bg']};
    color: {P['text2']};
    font-size: 10pt;
    font-weight: bold;
    padding: 10px 22px;
    border: none;
    border-bottom: 3px solid transparent;
    min-width: 130px;
}}
QTabBar::tab:selected {{
    color: {P['white']};
    border-bottom: 3px solid {P['red']};
    background: {P['bg2']};
}}
QTabBar::tab:hover:!selected {{
    color: {P['text']};
    background: {P['bg3']};
}}
QPushButton {{
    background-color: {P['red']};
    color: {P['white']};
    border: none;
    border-radius: 4px;
    padding: 7px 18px;
    font-size: 10pt;
    font-weight: bold;
}}
QPushButton:hover {{ background-color: {P['red2']}; }}
QPushButton:pressed {{ background-color: #A00015; }}
QPushButton[secondary="true"] {{
    background-color: {P['bg4']};
    color: {P['text']};
    border: 1px solid {P['border']};
}}
QPushButton[secondary="true"]:hover {{ background-color: {P['bg3']}; border-color: {P['red']}; }}
QPushButton[danger="true"] {{
    background-color: transparent;
    color: {P['red']};
    border: 1px solid {P['red']};
}}
QPushButton[danger="true"]:hover {{ background-color: {P['red']}; color: white; }}
QLineEdit, QSpinBox, QComboBox {{
    background-color: {P['bg4']};
    color: {P['text']};
    border: 1px solid {P['border']};
    border-radius: 4px;
    padding: 5px 8px;
    font-size: 10pt;
    selection-background-color: {P['red']};
}}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus {{
    border: 1px solid {P['red']};
}}
QComboBox::drop-down {{ border: none; width: 20px; }}
QComboBox QAbstractItemView {{
    background: {P['bg3']};
    color: {P['text']};
    selection-background-color: {P['red']};
    border: 1px solid {P['border']};
}}
QListWidget {{
    background-color: {P['bg4']};
    color: {P['text']};
    border: 1px solid {P['border']};
    border-radius: 4px;
    outline: none;
    font-size: 10pt;
}}
QListWidget::item {{ padding: 5px 8px; border-radius: 3px; }}
QListWidget::item:selected {{ background: {P['red']}; color: white; }}
QListWidget::item:hover:!selected {{ background: {P['bg3']}; }}
QScrollBar:vertical {{
    background: {P['bg']};
    width: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:vertical {{
    background: {P['border']};
    border-radius: 4px;
    min-height: 30px;
}}
QScrollBar::handle:vertical:hover {{ background: {P['red']}; }}
QScrollBar::add-line, QScrollBar::sub-line {{ height: 0; }}
QScrollBar:horizontal {{
    background: {P['bg']};
    height: 8px;
    border-radius: 4px;
}}
QScrollBar::handle:horizontal {{
    background: {P['border']};
    border-radius: 4px;
    min-width: 30px;
}}
QScrollBar::handle:horizontal:hover {{ background: {P['red']}; }}
QGroupBox {{
    color: {P['red']};
    font-weight: bold;
    font-size: 10pt;
    border: 1px solid {P['border']};
    border-radius: 6px;
    margin-top: 12px;
    padding-top: 8px;
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 6px;
    color: {P['red']};
}}
QCheckBox, QRadioButton {{
    color: {P['text']};
    spacing: 8px;
}}
QCheckBox::indicator, QRadioButton::indicator {{
    width: 16px;
    height: 16px;
    border: 2px solid {P['border']};
    border-radius: 3px;
    background: {P['bg4']};
}}
QCheckBox::indicator:checked {{
    background: {P['red']};
    border-color: {P['red']};
}}
QRadioButton::indicator {{ border-radius: 8px; }}
QRadioButton::indicator:checked {{
    background: {P['red']};
    border-color: {P['red']};
}}
QTableWidget {{
    background: {P['bg3']};
    color: {P['text']};
    gridline-color: {P['border']};
    border: 1px solid {P['border']};
    border-radius: 4px;
    outline: none;
}}
QTableWidget::item {{ padding: 4px 8px; }}
QTableWidget::item:selected {{ background: {P['red']}; color: white; }}
QHeaderView::section {{
    background: {P['bg']};
    color: {P['text2']};
    font-weight: bold;
    padding: 6px 8px;
    border: none;
    border-right: 1px solid {P['border']};
    border-bottom: 1px solid {P['border']};
}}
QHeaderView::section:first {{ border-left: none; }}
QSplitter::handle {{ background: {P['border']}; width: 1px; }}
QMessageBox {{ background: {P['bg2']}; }}
QDialog {{ background: {P['bg2']}; }}
QLabel#dim {{ color: {P['text2']}; font-size: 9pt; }}
QFrame#card {{
    background: {P['bg3']};
    border: 1px solid {P['border']};
    border-radius: 8px;
}}
QFrame#divider {{
    background: {P['border']};
    max-height: 1px;
}}
"""

# ═══════════════════════════════════════════════════════════════
#  Daten-Modell (identisch mit tkinter-Version)
# ═══════════════════════════════════════════════════════════════

class TeamDaten:
    def __init__(self):
        self.mitarbeiter = []
        self.temporaere_mas = []
        self.takte = []
        self.qualifikationen = {}
        self.abwesenheiten = {}
        self.archiv = []
        self.gesamt_zaehler = {}
        self.takt_besetzung = {}
        self.takt_farben = {}
        self.rotations_modus = False
        self.rotations_reihenfolge = []
        self.rotations_start = {}
        self.fixierungen = {}  # {takt: [ma_primaer, ma_vertreter1, ...]}

    @property
    def alle_mas(self):
        """Gibt alle MAs zurueck. Gecacht solange sich Listen nicht aendern."""
        return list(self.mitarbeiter) + list(self.temporaere_mas)

    def alle_mas_set(self):
        """Als Set fuer schnelle Membership-Checks."""
        return set(self.mitarbeiter) | set(self.temporaere_mas)

    def ist_temporaer(self, ma):
        return ma in self.temporaere_mas

    def ma_kann_takt(self, ma, takt):
        return takt in self.qualifikationen.get(ma, set())

    def get_besetzung(self, takt):
        return self.takt_besetzung.get(takt, 1)

    def gesamt_zaehler_aktualisieren(self, plan):
        for tag_plan in plan.values():
            for ma, takt in tag_plan.items():
                if takt and takt != "ABWESEND":
                    if ma not in self.gesamt_zaehler:
                        self.gesamt_zaehler[ma] = {}
                    self.gesamt_zaehler[ma][takt] = self.gesamt_zaehler[ma].get(takt, 0) + 1

    def gesamt_zaehler_fuer_ma(self, ma):
        return self.gesamt_zaehler.get(ma, {})

    def takt_farbe(self, takt):
        idx = self.takte.index(takt) if takt in self.takte else 0
        return self.takt_farben.get(takt, TAKT_PRESET[idx % len(TAKT_PRESET)])

    def speichern(self, pfad=SAVE_FILE):
        data = {
            "config_version": CONFIG_VERSION,
            "mitarbeiter": self.mitarbeiter,
            "temporaere_mas": self.temporaere_mas,
            "takte": self.takte,
            "qualifikationen": {k: list(v) for k, v in self.qualifikationen.items()},
            "takt_besetzung": self.takt_besetzung,
            "takt_farben": self.takt_farben,
            "rotations_modus": self.rotations_modus,
            "rotations_reihenfolge": self.rotations_reihenfolge,
            "rotations_start": self.rotations_start,
            "fixierungen": self.fixierungen,
            "abwesenheiten": {k: list(v) for k, v in self.abwesenheiten.items()},
            "archiv": self.archiv,
            "gesamt_zaehler": self.gesamt_zaehler,
        }
        with open(pfad, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def laden(self, pfad=SAVE_FILE):
        # Backup vor dem Laden erstellen
        try:
            import shutil
            shutil.copy2(pfad, pfad + ".backup")
        except Exception:
            pass
        with open(pfad, "r", encoding="utf-8") as f:
            raw = f.read().strip()
        if not raw:
            raise ValueError("Konfigurationsdatei ist leer.")
        data = json.loads(raw)
        if not isinstance(data, dict):
            raise ValueError("Ungueltiges Dateiformat.")
        # Versions-Check: bei aelterer Version trotzdem laden aber migrieren
        datei_version = data.get("config_version", 1)
        if datei_version > CONFIG_VERSION:
            raise ValueError(
                f"Config-Version {datei_version} ist neuer als "
                f"diese Programmversion ({CONFIG_VERSION}). "
                "Bitte Programm aktualisieren.")
        self.mitarbeiter = data.get("mitarbeiter", [])
        self.temporaere_mas = data.get("temporaere_mas", [])
        self.takte = data.get("takte", [])
        self.qualifikationen = {k: set(v) for k, v in data.get("qualifikationen", {}).items()}
        self.takt_besetzung = data.get("takt_besetzung", {})
        self.takt_farben = data.get("takt_farben", {})
        self.rotations_modus = data.get("rotations_modus", False)
        self.rotations_reihenfolge = data.get("rotations_reihenfolge", [])
        self.rotations_start = data.get("rotations_start", {})
        raw_fix = data.get("fixierungen", {})
        self.fixierungen = {}
        for t, v in raw_fix.items():
            self.fixierungen[t] = v if isinstance(v, list) else [v]
        self.abwesenheiten = {k: set(v) for k, v in data.get("abwesenheiten", {}).items()}
        self.archiv = data.get("archiv", [])
        self.gesamt_zaehler = data.get("gesamt_zaehler", {})

    def backup_wiederherstellen(self):
        backup = SAVE_FILE + ".backup"
        if os.path.exists(backup):
            import shutil
            shutil.copy2(backup, SAVE_FILE)
            self.laden()
            return True
        return False


# ═══════════════════════════════════════════════════════════════
#  Planungs-Logik
# ═══════════════════════════════════════════════════════════════

def plan_rotation(daten, tage):
    reihenfolge = daten.rotations_reihenfolge or list(daten.alle_mas)
    # Ketten auflösen: erster verfügbarer MA in der Kette bekommt den Takt
    def kette_auflösen(takt, tag):
        kette = daten.fixierungen.get(takt, [])
        if isinstance(kette, str):
            kette = [kette]
        abw = daten.abwesenheiten.get(tag, set())
        for ma in kette:
            if ma not in abw:
                return ma
        return None  # Alle abwesend -> Takt leer

    fixierte_takte = set(daten.fixierungen.keys())

    # Nur Takte die NICHT fixiert sind rotieren
    rotierende_takte = [t for t in daten.takte if t not in fixierte_takte]

    if not rotierende_takte:
        rotierende_takte = daten.takte  # Fallback falls alle fixiert

    # Startpunkt: entweder manuell gesetzt oder zufaellig (wechselt jede Woche)
    erster_ma = reihenfolge[0] if reihenfolge else None
    if erster_ma and erster_ma in daten.rotations_start:
        # Manuell gesetzter Startpunkt hat Vorrang
        start_takt = daten.rotations_start[erster_ma]
        start_idx = rotierende_takte.index(start_takt) if start_takt in rotierende_takte else 0
    else:
        # Historisch informierter Startpunkt:
        # Waehle den Takt den erster_ma zuletzt am wenigsten hatte
        # Damit rotiert der Startpunkt sinnvoll von Woche zu Woche
        if erster_ma and rotierende_takte:
            hist_erster = daten.gesamt_zaehler.get(erster_ma, {})
            # Takt mit niedrigstem historischem Zaehler = bevorzugter Start
            start_takt = min(rotierende_takte,
                key=lambda t: (hist_erster.get(t, 0), random.random()))
            start_idx = rotierende_takte.index(start_takt)
        else:
            start_idx = random.randint(0, max(len(rotierende_takte) - 1, 0))

    plan = {}
    for ti, tag in enumerate(tage):
        plan[tag] = {}
        abw = daten.abwesenheiten.get(tag, set())

        # Abwesende markieren
        for ma in daten.alle_mas:
            if ma in abw:
                plan[tag][ma] = "ABWESEND"

        # Fixierte Takte per Kette auflösen
        fixiert_heute = {}  # {ma: takt} für diesen Tag
        for takt in fixierte_takte:
            ma = kette_auflösen(takt, tag)
            if ma:
                plan[tag][ma] = takt
                fixiert_heute[ma] = takt

        # Offset pro Tag: jeden Tag einen Schritt weiter
        offset = (start_idx + ti) % len(rotierende_takte)

        # Nur nicht-fixierte, nicht-abwesende MAs rotieren
        rotierende_mas = [ma for ma in reihenfolge
                          if ma not in abw and ma not in fixiert_heute]

        for i, ma in enumerate(rotierende_mas):
            takt_idx = (offset + i) % len(rotierende_takte)
            plan[tag][ma] = rotierende_takte[takt_idx]

    return plan


def plan_erstellen(daten, tage):
    """
    Target-basierter Algorithmus.

    Schritt 1: Berechne faires Wochenziel (target) fuer jeden MA/Takt.
      - Basiert auf Qualifikationen und Anwesenheitstagen
      - Wird durch historische Daten (letzte Wochen) angepasst:
        wer zuletzt viel auf einem Takt war bekommt ein niedrigeres Ziel

    Schritt 2: Tagesweise Zuteilung per "groesste Abweichung vom Ziel"
      - Pro Tag: jeden Slot dem MA zuweisen dessen Ist-Wert am weitesten
        unter seinem Ziel liegt (groesste Schulden)
      - Qualifikationen werden hart geprueft
      - Besetzungszahl wird pro Tag genau eingehalten
    """
    def kette_fuer_tag(takt, tag):
        kette = daten.fixierungen.get(takt, [])
        if isinstance(kette, str): kette = [kette]
        abw = daten.abwesenheiten.get(tag, set())
        for ma in kette:
            if ma not in abw: return ma
        return None

    def ma_kann(ma, takt):
        """Prueft ob MA diesen Takt ausfuehren kann.
        Leere Qualifikations-Menge = MA kann alle Takte (z.B. neu angelegter MA).
        Explizit gesetzte Qualifikationen = nur diese Takte erlaubt."""
        quali = daten.qualifikationen.get(ma, set())
        return not quali or takt in quali

    fixierte_takte  = set(daten.fixierungen.keys())
    rot_takte       = [t for t in daten.takte if t not in fixierte_takte]
    n_tage          = len(tage)

    # ── Schritt 1: Ziel-Matrix ──────────────────────────────────────
    # Historische Belastung: direkt aus gesamt_zaehler lesen (einzige Wahrheitsquelle)
    # gesamt_zaehler wird nur beim finalen Export/Archivieren aktualisiert
    # -> kein Auseinanderlaufen moeglich
    hist_last = {ma: dict(daten.gesamt_zaehler.get(ma, {}))
                 for ma in daten.alle_mas}

    # Anwesenheitstage pro MA diese Woche
    anw = {ma: sum(1 for t in tage
                   if ma not in daten.abwesenheiten.get(t, set()))
           for ma in daten.alle_mas}

    # Fixierte MAs bestimmen
    fixiert_ma_gesamt = set()
    for kette in daten.fixierungen.values():
        k = kette if isinstance(kette, list) else [kette]
        if k: fixiert_ma_gesamt.add(k[0])

    rot_mas = [m for m in daten.alle_mas if m not in fixiert_ma_gesamt]

    # Verfuegbare (nicht-fixierte, anwesende) MAs pro Tag
    verf_pro_tag = {
        tag: [m for m in rot_mas
              if m not in daten.abwesenheiten.get(tag, set())]
        for tag in tage
    }

    # Gesamt-Slots pro Takt ueber die Woche
    slots = {t: daten.get_besetzung(t) * n_tage for t in rot_takte}

    # Qualifizierte MAs pro Takt (unter den rot_mas)
    quali_mas = {t: [m for m in rot_mas if ma_kann(m, t)] for t in rot_takte}

    # Rohes Ziel: slots / anzahl_qualifizierter_mas * anwesenheitstage_ma
    # = proportionaler Anteil den dieser MA an diesem Takt haben sollte
    target_roh = {}
    for ma in rot_mas:
        target_roh[ma] = {}
        for takt in rot_takte:
            qmas = [m for m in quali_mas[takt] if anw[m] > 0]
            if not qmas or ma not in qmas:
                target_roh[ma][takt] = 0.0
                continue
            # Proportionaler Anteil basierend auf Anwesenheitstagen
            gesamt_anw = sum(anw[m] for m in qmas)
            target_roh[ma][takt] = slots[takt] * anw[ma] / max(gesamt_anw, 1)

    # Ueberhang-Ziel
    plaetze_pro_tag = {
        tag: sum(daten.get_besetzung(t) for t in rot_takte)
        for tag in tage
    }
    uh_slots = sum(
        max(0, len(verf_pro_tag[tag]) - plaetze_pro_tag[tag])
        for tag in tage
    )
    for ma in rot_mas:
        if anw[ma] > 0:
            gesamt_anw = sum(anw[m] for m in rot_mas if anw[m] > 0)
            target_roh[ma]["UEBERHANG"] = uh_slots * anw[ma] / max(gesamt_anw, 1)
        else:
            target_roh[ma]["UEBERHANG"] = 0.0

    # Historische Anpassung: wer in letzten Wochen viel hatte bekommt
    # einen Abzug proportional zur Ueberbelastung
    target = {}
    for ma in rot_mas:
        target[ma] = {}
        for takt in rot_takte + ["UEBERHANG"]:
            roh = target_roh[ma].get(takt, 0.0)
            hist_val = hist_last.get(ma, {}).get(takt, 0)
            # Erwarteter historischer Wert (wenn alles fair waere)
            n_hist_wochen = max(len(hist_wochen), 1)
            erwartet = roh * n_hist_wochen
            # Anpassungsfaktor: wer mehr als erwartet hatte bekommt Abzug
            ueber = hist_val - erwartet
            anpassung = ueber * 0.3  # 30% Anpassung pro Woche
            target[ma][takt] = max(0.0, roh - anpassung)

    # ── Schritt 2: Tagesweise Zuteilung ─────────────────────────────
    # Ist-Zaehler: wie oft hatte MA diesen Takt DIESE Woche
    ist = {ma: defaultdict(float) for ma in rot_mas}

    def schulden(ma, takt):
        """Wie weit ist MA unter seinem Ziel? Groesser = bevorzugter."""
        return target[ma].get(takt, 0.0) - ist[ma][takt]

    plan = {tag: {} for tag in tage}

    for tag in tage:
        abw = daten.abwesenheiten.get(tag, set())

        # Abwesende markieren
        for ma in daten.alle_mas:
            if ma in abw:
                plan[tag][ma] = "ABWESEND"

        # Fixierte Takte per Kette besetzen
        fixiert_heute = {}
        for takt in fixierte_takte:
            ma = kette_fuer_tag(takt, tag)
            if ma and ma not in abw:
                plan[tag][ma] = takt
                fixiert_heute[ma] = takt

        # Verfuegbare rotierende MAs heute
        verf_heute = [m for m in rot_mas if m not in abw and m not in fixiert_heute]

        # Taktplaetze und Ueberhang berechnen
        plaetze = sum(daten.get_besetzung(t) for t in rot_takte)
        n_verf  = len(verf_heute)
        n_uh    = max(0, n_verf - plaetze)

        # ── Takt-Zuteilung ZUERST (oberste Prioritaet) ─────────────────
        # Erst alle Taktplaetze besetzen, dann Ueberhang vergeben
        takt_vergaben = defaultdict(int)

        # MAs mit wenigsten Qualifikationen zuerst einteilen
        nicht_zu = sorted(verf_heute,
            key=lambda m: (sum(1 for t in rot_takte if ma_kann(m, t)), random.random()))

        zugewiesen = set()

        while nicht_zu:
            # Pruefe ob noch Taktplaetze frei sind
            noch_frei = any(takt_vergaben[t] < daten.get_besetzung(t) for t in rot_takte)

            if not noch_frei:
                # Alle Taktplaetze besetzt -> restliche MAs bekommen Ueberhang
                for ma in nicht_zu:
                    plan[tag][ma] = "UEBERHANG"
                    ist[ma]["UEBERHANG"] += 1
                break

            # Finde bestes MA-Takt-Paar mit groessten Schulden
            bestes_paar  = None
            beste_schuld = -999999

            for ma in nicht_zu:
                frei = [t for t in rot_takte
                        if ma_kann(ma, t)
                        and takt_vergaben[t] < daten.get_besetzung(t)]
                if not frei:
                    continue
                bester_takt = max(frei, key=lambda t: (schulden(ma, t), random.random()))
                s = schulden(ma, bester_takt)
                if s > beste_schuld:
                    beste_schuld = s
                    bestes_paar  = (ma, bester_takt)

            if bestes_paar is None:
                # Kein MA mehr qualifiziert fuer freie Takte
                # -> Forciere Besetzung: nimm irgendeinen qualifizierten MA
                for takt in rot_takte:
                    while takt_vergaben[takt] < daten.get_besetzung(takt):
                        kandidaten = [m for m in nicht_zu if ma_kann(m, takt)]
                        if not kandidaten:
                            break  # wirklich niemand qualifiziert
                        ma = min(kandidaten,
                            key=lambda m: (ist[m][takt], hist_last.get(m, {}).get(takt, 0), random.random()))
                        plan[tag][ma] = takt
                        ist[ma][takt] += 1
                        takt_vergaben[takt] += 1
                        nicht_zu.remove(ma)
                # Verbleibende -> Ueberhang
                for ma in nicht_zu:
                    plan[tag][ma] = "UEBERHANG"
                    ist[ma]["UEBERHANG"] += 1
                break

            ma, takt = bestes_paar
            plan[tag][ma] = takt
            ist[ma][takt] += 1
            takt_vergaben[takt] += 1
            nicht_zu.remove(ma)

        # Sicherheits-Check: fehlende Taktplaetze auffuellen
        for takt in rot_takte:
            while takt_vergaben[takt] < daten.get_besetzung(takt):
                # Suche MA der noch keinen Takt hat und qualifiziert ist
                kandidaten = [m for m in verf_heute
                              if plan[tag].get(m) in ("UEBERHANG", None)
                              and ma_kann(m, takt)]
                if not kandidaten:
                    break
                ma = min(kandidaten,
                    key=lambda m: (ist[m][takt], hist_last.get(m, {}).get(takt, 0), random.random()))
                plan[tag][ma] = takt
                ist[ma][takt] += 1
                takt_vergaben[takt] += 1

    return plan

# ═══════════════════════════════════════════════════════════════
#  UI Helper Widgets
# ═══════════════════════════════════════════════════════════════

def card(parent=None):
    f = QFrame(parent)
    f.setObjectName("card")
    f.setLayout(QVBoxLayout())
    f.layout().setContentsMargins(16, 16, 16, 16)
    f.layout().setSpacing(10)
    return f

def divider(parent=None):
    f = QFrame(parent)
    f.setObjectName("divider")
    f.setFixedHeight(1)
    return f

def label(text, bold=False, dim=False, size=10, parent=None):
    l = QLabel(text, parent)
    f = QFont("Segoe UI", size)
    f.setBold(bold)
    l.setFont(f)
    if dim:
        l.setObjectName("dim")
        l.setStyleSheet(f"color: {P['text2']};")
    return l

def sec_btn(text, parent=None):
    b = QPushButton(text, parent)
    b.setProperty("secondary", True)
    b.style().polish(b)
    return b

def danger_btn(text, parent=None):
    b = QPushButton(text, parent)
    b.setProperty("danger", True)
    b.style().polish(b)
    return b


# ═══════════════════════════════════════════════════════════════
#  Header Widget
# ═══════════════════════════════════════════════════════════════

class HeaderWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setFixedHeight(64)
        self.setStyleSheet(f"background: {P['bg']};")
        lay = QHBoxLayout(self)
        lay.setContentsMargins(28, 0, 28, 0)

        # Clicktakt Schriftzug
        brand = QLabel("Clicktakt")
        brand.setStyleSheet(
            "color: #FFFFFF;"
            "font-size: 26pt;"
            "font-weight: 300;"
            "font-family: 'Segoe UI Light', 'Helvetica Neue', 'Arial', sans-serif;"
            "letter-spacing: 3px;"
        )
        lay.addWidget(brand)

        # Vertikaler Trenner
        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet(f"color: {P['border']}; margin: 16px 0;")
        lay.addWidget(sep)
        lay.addSpacing(12)

        # Untertitel
        sub = QLabel("Werk Leipzig")
        sub.setStyleSheet(
            f"color: {P['text2']};"
            "font-size: 10pt;"
            "letter-spacing: 1px;"
        )
        lay.addWidget(sub)
        lay.addStretch()

        # Status
        st = QLabel("●  AKTIV")
        st.setStyleSheet(f"color:{P['green']};font-size:9pt;font-weight:bold;letter-spacing:1px;")
        lay.addWidget(st)

        # Trennlinie unten
        line = QFrame(self)
        line.setObjectName("divider")
        line.setGeometry(0, 63, 9999, 1)

    def resizeEvent(self, e):
        super().resizeEvent(e)
        for c in self.children():
            if isinstance(c, QFrame) and c.objectName() == "divider":
                c.setGeometry(0, 63, self.width(), 1)



# ═══════════════════════════════════════════════════════════════
#  Tab 1: Stammdaten
# ═══════════════════════════════════════════════════════════════

class StammdatenTab(QWidget):
    changed = pyqtSignal()

    def __init__(self, daten):
        super().__init__()
        self.daten = daten
        self._build()

    def _build(self):
        main = QHBoxLayout(self)
        main.setContentsMargins(16, 16, 16, 16)
        main.setSpacing(12)

        # ── Linke Spalte: Stammteam + Temporaer ──
        left = QVBoxLayout()
        left.setSpacing(12)

        # Stammteam
        g_stamm = QGroupBox("Stammteam")
        vl = QVBoxLayout(g_stamm)
        inp = QHBoxLayout()
        self.entry_ma = QLineEdit()
        self.entry_ma.setPlaceholderText("Name eingeben...")
        self.entry_ma.returnPressed.connect(self._ma_add)
        btn_add = QPushButton("+")
        btn_add.setFixedWidth(40)
        btn_add.clicked.connect(self._ma_add)
        inp.addWidget(self.entry_ma)
        inp.addWidget(btn_add)
        vl.addLayout(inp)
        self.lb_ma = QListWidget()
        self.lb_ma.setMinimumHeight(180)
        self.lb_ma.keyPressEvent = lambda e: (self._ma_del() if e.key() == Qt.Key.Key_Delete else QListWidget.keyPressEvent(self.lb_ma, e))
        vl.addWidget(self.lb_ma)
        btn_del = danger_btn("Entfernen")
        btn_del.clicked.connect(self._ma_del)
        vl.addWidget(btn_del)
        left.addWidget(g_stamm, 2)

        # Temporaer
        g_temp = QGroupBox("Temporär  (T)")
        g_temp.setStyleSheet(f"QGroupBox {{ color: {P['gold']}; }}"
                              f"QGroupBox::title {{ color: {P['gold']}; }}")
        vl2 = QVBoxLayout(g_temp)
        inp2 = QHBoxLayout()
        self.entry_temp = QLineEdit()
        self.entry_temp.setPlaceholderText("Name eingeben...")
        self.entry_temp.returnPressed.connect(self._temp_add)
        btn_add2 = QPushButton("+")
        btn_add2.setFixedWidth(40)
        btn_add2.clicked.connect(self._temp_add)
        btn_del2 = danger_btn("Entfernen")
        btn_del2.clicked.connect(self._temp_del)
        inp2.addWidget(self.entry_temp)
        inp2.addWidget(btn_add2)
        inp2.addWidget(btn_del2)
        vl2.addLayout(inp2)
        self.lb_temp = QListWidget()
        self.lb_temp.setMaximumHeight(110)
        self.lb_temp.setStyleSheet(f"background:{P['bg4']};color:{P['gold']};")
        self.lb_temp.keyPressEvent = lambda e: (self._temp_del() if e.key() == Qt.Key.Key_Delete else QListWidget.keyPressEvent(self.lb_temp, e))
        vl2.addWidget(self.lb_temp)
        lbl_hint = label("Qualifikationen werden beim Hinzufügen abgefragt", dim=True)
        vl2.addWidget(lbl_hint)
        left.addWidget(g_temp, 1)

        main.addLayout(left, 1)

        # ── Mittlere Spalte: Takte ──
        g_takte = QGroupBox("Takte  (× = Plätze pro Tag)")
        vl3 = QVBoxLayout(g_takte)
        inp3 = QHBoxLayout()
        self.entry_takt = QLineEdit()
        self.entry_takt.setPlaceholderText("Taktname...")
        self.entry_takt.returnPressed.connect(self._takt_add)
        lbl_x = label("×")
        self.spin_bes = QSpinBox()
        self.spin_bes.setRange(1, 20)
        self.spin_bes.setFixedWidth(55)
        btn_add3 = QPushButton("+")
        btn_add3.setFixedWidth(40)
        btn_add3.clicked.connect(self._takt_add)
        inp3.addWidget(self.entry_takt)
        inp3.addWidget(lbl_x)
        inp3.addWidget(self.spin_bes)
        inp3.addWidget(btn_add3)
        vl3.addLayout(inp3)
        self.lb_takt = QListWidget()
        self.lb_takt.keyPressEvent = lambda e: (self._takt_del() if e.key() == Qt.Key.Key_Delete else QListWidget.keyPressEvent(self.lb_takt, e))
        vl3.addWidget(self.lb_takt)
        btn_row = QHBoxLayout()
        btn_bes = sec_btn("Besetzung ändern")
        btn_bes.clicked.connect(self._takt_bes_aendern)
        btn_del3 = danger_btn("Entfernen")
        btn_del3.clicked.connect(self._takt_del)
        btn_row.addWidget(btn_bes)
        btn_row.addWidget(btn_del3)
        vl3.addLayout(btn_row)
        main.addWidget(g_takte, 1)

        # ── Rechte Spalte: Fixierungen ──
        g_fix = QGroupBox("Fixierungen  –  Vertreterkette")
        vl4 = QVBoxLayout(g_fix)

        # Takt-Auswahl
        inp4 = QHBoxLayout()
        lbl_ft = label("Takt:")
        self.cb_fix_takt = QComboBox()
        btn_fix = QPushButton("Kette festlegen")
        btn_fix.clicked.connect(self._fix_setzen)
        inp4.addWidget(lbl_ft)
        inp4.addWidget(self.cb_fix_takt)
        inp4.addWidget(btn_fix)
        vl4.addLayout(inp4)

        vl4.addWidget(label("Klicke auf einen Eintrag zum Bearbeiten", dim=True))
        self.lb_fix = QListWidget()
        self.lb_fix.setMaximumHeight(140)
        self.lb_fix.itemDoubleClicked.connect(self._fix_edit)
        self.lb_fix.keyPressEvent = lambda e: (self._fix_del() if e.key() == Qt.Key.Key_Delete else QListWidget.keyPressEvent(self.lb_fix, e))
        vl4.addWidget(self.lb_fix)
        btn_fix_del = danger_btn("Fixierung entfernen")
        btn_fix_del.clicked.connect(self._fix_del)
        vl4.addWidget(btn_fix_del)
        vl4.addStretch()

        # Buttons unten
        vl4.addWidget(divider())
        btn_row2 = QHBoxLayout()
        btn_save = sec_btn("Konfiguration speichern")
        btn_save.clicked.connect(self._speichern)
        btn_load = sec_btn("Konfiguration laden")
        btn_load.clicked.connect(self._laden)
        btn_row2.addWidget(btn_save)
        btn_row2.addWidget(btn_load)
        vl4.addLayout(btn_row2)

        main.addWidget(g_fix, 1)

        self.refresh()

    def refresh(self):
        self.lb_ma.clear()
        for m in self.daten.mitarbeiter:
            self.lb_ma.addItem(m)
        self.lb_temp.clear()
        for m in self.daten.temporaere_mas:
            it = QListWidgetItem(m)
            it.setForeground(QColor(P['gold']))
            self.lb_temp.addItem(it)
        self.lb_takt.clear()
        for t in self.daten.takte:
            n = self.daten.get_besetzung(t)
            txt = f"{t}  ×{n}" if n > 1 else t
            self.lb_takt.addItem(txt)
        self.cb_fix_takt.clear()
        self.cb_fix_takt.addItems(self.daten.takte)
        self.lb_fix.clear()
        for t, kette in self.daten.fixierungen.items():
            kette_str = " → ".join(kette) if isinstance(kette, list) else kette
            self.lb_fix.addItem(f"{t}  :  {kette_str}")

    def _ma_add(self):
        name = self.entry_ma.text().strip()
        if name and name not in self.daten.alle_mas_set():
            self.daten.mitarbeiter.append(name)
            self.daten.qualifikationen[name] = set()
            self.entry_ma.clear()
            self.refresh()
            self.changed.emit()

    def _ma_del(self):
        sel = self.lb_ma.currentItem()
        if not sel: return
        name = sel.text()
        self.daten.mitarbeiter.remove(name)
        self.daten.qualifikationen.pop(name, None)
        self.refresh()
        self.changed.emit()

    def _temp_add(self):
        name = self.entry_temp.text().strip()
        if not name: return
        if name in self.daten.alle_mas_set():
            QMessageBox.warning(self, "Bereits vorhanden", f"'{name}' existiert bereits.")
            return
        self.entry_temp.clear()
        self.daten.temporaere_mas.append(name)
        self.daten.qualifikationen[name] = set()
        self.refresh()
        self.changed.emit()
        # Qualifikations-Popup
        dlg = TempQualiDialog(name, self.daten.takte, self)
        if dlg.exec():
            self.daten.qualifikationen[name] = set(dlg.selected)
            self.changed.emit()

    def _temp_del(self):
        sel = self.lb_temp.currentItem()
        if not sel: return
        name = sel.text()
        self.daten.temporaere_mas.remove(name)
        self.daten.qualifikationen.pop(name, None)
        self.refresh()
        self.changed.emit()

    def _takt_add(self):
        name = self.entry_takt.text().strip()
        if name and name not in self.daten.takte:
            self.daten.takte.append(name)
            self.daten.takt_besetzung[name] = self.spin_bes.value()
            self.entry_takt.clear()
            self.spin_bes.setValue(1)
            self.refresh()
            self.changed.emit()

    def _takt_del(self):
        sel = self.lb_takt.currentItem()
        if not sel: return
        name = sel.text().split("  ×")[0].strip()
        self.daten.takte.remove(name)
        self.daten.takt_besetzung.pop(name, None)
        self.daten.fixierungen.pop(name, None)
        for ma in self.daten.alle_mas:
            self.daten.qualifikationen.get(ma, set()).discard(name)
        self.refresh()
        self.changed.emit()

    def _takt_bes_aendern(self):
        sel = self.lb_takt.currentItem()
        if not sel: return
        name = sel.text().split("  ×")[0].strip()
        self.daten.takt_besetzung[name] = self.spin_bes.value()
        self.refresh()

    def _fix_setzen(self, takt_vorgabe=None):
        takt = takt_vorgabe or self.cb_fix_takt.currentText()
        if not takt:
            return
        bestehend = self.daten.fixierungen.get(takt, [])
        if isinstance(bestehend, str):
            bestehend = [bestehend]
        dlg = FixKetteDialog(takt, self.daten.alle_mas, bestehend, self)
        if dlg.exec():
            if dlg.kette:
                self.daten.fixierungen[takt] = dlg.kette
            else:
                self.daten.fixierungen.pop(takt, None)
            self.refresh()

    def _fix_edit(self, item):
        takt = item.text().split("  :")[0].strip()
        self._fix_setzen(takt_vorgabe=takt)

    def _fix_del(self):
        sel = self.lb_fix.currentItem()
        if not sel: return
        takt = sel.text().split("  :")[0].strip()
        self.daten.fixierungen.pop(takt, None)
        self.refresh()

    def _speichern(self):
        self.daten.speichern()
        QMessageBox.information(self, "Gespeichert", "Konfiguration gespeichert.")

    def _laden(self):
        pfad, _ = QFileDialog.getOpenFileName(self, "Laden", _BASIS, "JSON (*.json)")
        if pfad:
            self.daten.laden(pfad)
            self.refresh()
            self.changed.emit()


class FixKetteDialog(QDialog):
    """Dialog zum Festlegen einer Vertreterkette fuer einen Takt."""
    def __init__(self, takt, alle_mas, bestehend, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Vertreterkette: {takt}")
        self.setMinimumWidth(400)
        self.kette = list(bestehend)
        lay = QVBoxLayout(self)

        lay.addWidget(label(f"Vertreterkette für: {takt}", bold=True, size=11))
        lay.addWidget(label("Primaer -> Vertreter 1 -> ... | Alle abwesend = Takt leer", dim=True))
        lay.addWidget(divider())

        # Aktuelle Kette anzeigen
        lay.addWidget(label("Aktuelle Kette:", bold=True))
        self.lb_kette = QListWidget()
        self.lb_kette.setMaximumHeight(140)
        self._refresh_kette()
        lay.addWidget(self.lb_kette)

        # Buttons zum Bearbeiten
        btn_row = QHBoxLayout()
        btn_hoch = sec_btn("▲  Hoch")
        btn_run  = sec_btn("▼  Runter")
        btn_del  = danger_btn("Entfernen")
        btn_hoch.setFixedWidth(90)
        btn_run.setFixedWidth(90)
        btn_hoch.clicked.connect(self._hoch)
        btn_run.clicked.connect(self._runter)
        btn_del.clicked.connect(self._entfernen)
        btn_row.addWidget(btn_hoch)
        btn_row.addWidget(btn_run)
        btn_row.addStretch()
        btn_row.addWidget(btn_del)
        lay.addLayout(btn_row)

        lay.addWidget(divider())

        # MA hinzufügen
        lay.addWidget(label("MA zur Kette hinzufügen:", bold=True))
        add_row = QHBoxLayout()
        self.cb_add = QComboBox()
        verfuegbar = [m for m in alle_mas if m not in self.kette]
        self.cb_add.addItems(verfuegbar)
        btn_add = QPushButton("+ Hinzufügen")
        btn_add.clicked.connect(self._hinzufuegen)
        add_row.addWidget(self.cb_add)
        add_row.addWidget(btn_add)
        lay.addLayout(add_row)

        lay.addWidget(divider())

        # Bestätigung
        btn_ok_row = QHBoxLayout()
        btn_abbruch = sec_btn("Abbrechen")
        btn_ok = QPushButton("✓  Kette speichern")
        btn_abbruch.clicked.connect(self.reject)
        btn_ok.clicked.connect(self.accept)
        btn_ok_row.addWidget(btn_abbruch)
        btn_ok_row.addStretch()
        btn_ok_row.addWidget(btn_ok)
        lay.addLayout(btn_ok_row)

        self._alle_mas = alle_mas

    def _refresh_kette(self):
        self.lb_kette.clear()
        for i, ma in enumerate(self.kette):
            prefix = "👑 Primär:     " if i == 0 else f"   Vertreter {i}: "
            self.lb_kette.addItem(f"{prefix}{ma}")

    def _refresh_cb(self):
        self.cb_add.clear()
        self.cb_add.addItems([m for m in self._alle_mas if m not in self.kette])

    def _hinzufuegen(self):
        ma = self.cb_add.currentText()
        if ma and ma not in self.kette:
            self.kette.append(ma)
            self._refresh_kette()
            self._refresh_cb()

    def _entfernen(self):
        i = self.lb_kette.currentRow()
        if i >= 0:
            self.kette.pop(i)
            self._refresh_kette()
            self._refresh_cb()

    def _hoch(self):
        i = self.lb_kette.currentRow()
        if i > 0:
            self.kette[i-1], self.kette[i] = self.kette[i], self.kette[i-1]
            self._refresh_kette()
            self.lb_kette.setCurrentRow(i-1)

    def _runter(self):
        i = self.lb_kette.currentRow()
        if i >= 0 and i < len(self.kette)-1:
            self.kette[i+1], self.kette[i] = self.kette[i], self.kette[i+1]
            self._refresh_kette()
            self.lb_kette.setCurrentRow(i+1)


class TempQualiDialog(QDialog):
    def __init__(self, name, takte, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Qualifikationen: {name}")
        self.setMinimumWidth(340)
        self.selected = []
        lay = QVBoxLayout(self)
        lay.addWidget(label(f"Welche Takte kann {name} ausführen?", bold=True))
        lay.addWidget(label("(Kann später in Tab Qualifikationen angepasst werden)", dim=True))
        lay.addWidget(divider())
        self.checks = {}
        grid = QGridLayout()
        for i, t in enumerate(takte):
            cb = QCheckBox(t)
            grid.addWidget(cb, i//2, i%2)
            self.checks[t] = cb
        lay.addLayout(grid)
        btn_alle = sec_btn("Alle auswählen")
        btn_alle.clicked.connect(lambda: [c.setChecked(True) for c in self.checks.values()])
        lay.addWidget(btn_alle)
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._ok)
        btns.rejected.connect(self.reject)
        lay.addWidget(btns)

    def _ok(self):
        self.selected = [t for t, c in self.checks.items() if c.isChecked()]
        self.accept()


# ═══════════════════════════════════════════════════════════════
#  Tab 2: Qualifikationen
# ═══════════════════════════════════════════════════════════════

class QualifikationenTab(QWidget):
    def __init__(self, daten):
        super().__init__()
        self.daten = daten
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 16)

        top = QHBoxLayout()
        top.addWidget(label("Haken = MA kann diesen Takt ausführen", bold=True))
        top.addStretch()
        hinweis = QLabel("ℹ  Keine Haken = kann alle Takte")
        hinweis.setStyleSheet(f"color:{P['gold']};font-size:9pt;padding:4px 8px;"
                               f"background:{P['bg4']};border-radius:4px;")
        top.addWidget(hinweis)
        btn_alle = QPushButton("Alle für alle Takte qualifizieren")
        btn_alle.clicked.connect(self._alle_quali)
        top.addWidget(btn_alle)
        lay.addLayout(top)
        lay.addWidget(divider())

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"QScrollArea {{ border: none; background: {P['bg2']}; }}")
        self.inner = QWidget()
        self.inner.setStyleSheet(f"background: {P['bg2']};")
        self.grid = QGridLayout(self.inner)
        self.grid.setSpacing(4)
        scroll.setWidget(self.inner)
        lay.addWidget(scroll)

        self.refresh()

    def refresh(self):
        # Grid leeren
        while self.grid.count():
            item = self.grid.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        if not self.daten.mitarbeiter and not self.daten.temporaere_mas:
            self.grid.addWidget(label("Bitte zuerst Mitarbeiter anlegen.", dim=True), 0, 0)
            return
        if not self.daten.takte:
            self.grid.addWidget(label("Bitte zuerst Takte anlegen.", dim=True), 0, 0)
            return

        # Header-Zeile
        hdr = QLabel("MA \\ Takt")
        hdr.setStyleSheet(f"color:{P['text2']};font-weight:bold;padding:4px 8px;")
        self.grid.addWidget(hdr, 0, 0)
        for ti, takt in enumerate(self.daten.takte):
            lbl = QLabel(takt)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet(f"color:{P['red']};font-weight:bold;padding:4px 8px;"
                               f"border-bottom:2px solid {P['red']};")
            self.grid.addWidget(lbl, 0, ti+1)

        n_takte = len(self.daten.takte)
        # Alle-Button Spalte Header
        alle_hdr = QLabel("Alle")
        alle_hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
        alle_hdr.setStyleSheet(f"color:{P['text2']};font-weight:bold;padding:4px 8px;")
        self.grid.addWidget(alle_hdr, 0, n_takte+1)

        self._checks = {}
        for ri, ma in enumerate(self.daten.alle_mas):
            row = ri + 1
            # MA-Name
            is_temp = self.daten.ist_temporaer(ma)
            ma_lbl = QLabel(f"{'▲ ' if is_temp else ''}{ma}")
            ma_lbl.setStyleSheet(
                f"color:{P['gold'] if is_temp else P['text']};font-weight:bold;"
                f"padding:4px 8px;background:{P['bg4']};border-radius:3px;")
            self.grid.addWidget(ma_lbl, row, 0)

            self._checks[ma] = {}
            for ti, takt in enumerate(self.daten.takte):
                checked = takt in self.daten.qualifikationen.get(ma, set())
                cb = QCheckBox("✓" if checked else "✗")
                cb.setChecked(checked)
                cb.setStyleSheet(
                    "margin-left:auto;margin-right:auto;"
                    f"color: {'#27AE60' if checked else '#888888'};"
                    "font-size: 11pt; font-weight: bold;"
                )
                def _on_toggle(state, c=cb):
                    ok = c.isChecked()
                    c.setText("✓" if ok else "✗")
                    col = "#27AE60" if ok else "#888888"
                    c.setStyleSheet(
                        f"margin-left:auto;margin-right:auto;"
                        f"color:{col};font-size:11pt;font-weight:bold;"
                    )
                cb.stateChanged.connect(_on_toggle)
                self.grid.addWidget(cb, row, ti+1)
                self._checks[ma][takt] = cb

            # Alle-Button
            btn = sec_btn("Alle")
            btn.setFixedWidth(60)
            btn.clicked.connect(lambda _, m=ma: self._alle_fuer_ma(m))
            self.grid.addWidget(btn, row, n_takte+1)

    def _alle_fuer_ma(self, ma):
        for cb in self._checks.get(ma, {}).values():
            cb.setChecked(True)

    def _alle_quali(self):
        for ma in self.daten.alle_mas:
            self.daten.qualifikationen[ma] = set(self.daten.takte)
        self.refresh()

    def speichern(self):
        for ma, takte_d in self._checks.items():
            self.daten.qualifikationen[ma] = {t for t, cb in takte_d.items() if cb.isChecked()}


# ═══════════════════════════════════════════════════════════════
#  Tab 3: Abwesenheiten
# ═══════════════════════════════════════════════════════════════

class AbwesenheitenTab(QWidget):
    def __init__(self, daten):
        super().__init__()
        self.daten = daten
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(12)

        # Arbeitstage
        g = QGroupBox("Arbeitstage dieser Woche")
        hl = QHBoxLayout(g)
        self.tag_checks = {}
        for i, tag in enumerate(WOCHENTAGE):
            cb = QCheckBox(tag)
            cb.setChecked(i < 5)
            cb.stateChanged.connect(self.refresh_tabelle)
            hl.addWidget(cb)
            self.tag_checks[tag] = cb
        hl.addStretch()
        lay.addWidget(g)

        # Abwesenheitstabelle
        g2 = QGroupBox("Anwesenheit  (Grün = da,  Rot = abwesend)")
        vl = QVBoxLayout(g2)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"border:none;background:{P['bg2']};")
        self.tbl_widget = QWidget()
        self.tbl_widget.setStyleSheet(f"background:{P['bg2']};")
        self.tbl_layout = QGridLayout(self.tbl_widget)
        self.tbl_layout.setSpacing(4)
        scroll.setWidget(self.tbl_widget)
        vl.addWidget(scroll)
        lay.addWidget(g2, 1)

        self.refresh_tabelle()

    def aktive_tage(self):
        return [t for t in WOCHENTAGE if self.tag_checks[t].isChecked()]

    def refresh_tabelle(self):
        while self.tbl_layout.count():
            item = self.tbl_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        tage = self.aktive_tage()
        if not tage or not self.daten.alle_mas:
            self.tbl_layout.addWidget(label("Bitte Mitarbeiter und Tage auswählen.", dim=True), 0, 0)
            return

        # Header
        hdr = QLabel("")
        hdr.setFixedWidth(120)
        self.tbl_layout.addWidget(hdr, 0, 0)
        for ti, tag in enumerate(tage):
            lbl = QLabel(tag)
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lbl.setStyleSheet(f"color:{P['red']};font-weight:bold;padding:4px;")
            lbl.setFixedWidth(90)
            self.tbl_layout.addWidget(lbl, 0, ti+1)
        # Ganze Woche Spalte
        gw_hdr = QLabel("Da/Abw.")
        gw_hdr.setAlignment(Qt.AlignmentFlag.AlignCenter)
        gw_hdr.setStyleSheet(f"color:{P['text2']};font-size:9pt;font-weight:bold;")
        gw_hdr.setFixedWidth(100)
        self.tbl_layout.addWidget(gw_hdr, 0, len(tage)+1)

        for ri, ma in enumerate(self.daten.alle_mas):
            row = ri + 1
            is_temp = self.daten.ist_temporaer(ma)
            lbl = QLabel(f"{'▲ ' if is_temp else ''}{ma}")
            lbl.setStyleSheet(
                f"color:{P['gold'] if is_temp else P['text']};font-weight:bold;"
                f"padding:4px 8px;background:{P['bg4']};border-radius:3px;")
            lbl.setFixedWidth(120)
            self.tbl_layout.addWidget(lbl, row, 0)

            for ti, tag in enumerate(tage):
                abw = ma in self.daten.abwesenheiten.get(tag, set())
                btn = QPushButton("Abw." if abw else "Da")
                btn.setFixedSize(88, 30)
                self._style_abw_btn(btn, abw)
                btn.clicked.connect(lambda _, m=ma, t=tag: self._toggle(m, t))
                self.tbl_layout.addWidget(btn, row, ti+1)

            # Ganze Woche
            alle_abw = all(ma in self.daten.abwesenheiten.get(t, set()) for t in tage)
            btn_w = QPushButton("Da/Abw.")
            btn_w.setFixedSize(100, 30)
            btn_w.setStyleSheet(
                f"background:{P['bg4']};color:{P['text2']};border:1px solid {P['border']};"
                f"border-radius:4px;font-size:9pt;")
            btn_w.clicked.connect(lambda _, m=ma, tl=tage: self._ganze_woche(m, tl))
            self.tbl_layout.addWidget(btn_w, row, len(tage)+1)

    def _style_abw_btn(self, btn, abw):
        if abw:
            btn.setStyleSheet(
                f"background:#6B0010;color:#FF6B7A;border:1px solid #D5001C;"
                f"border-radius:4px;font-weight:bold;font-size:9pt;")
        else:
            btn.setStyleSheet(
                f"background:#0A3D1A;color:#4CAF50;border:1px solid #27AE60;"
                f"border-radius:4px;font-weight:bold;font-size:9pt;")

    def _toggle(self, ma, tag):
        if tag not in self.daten.abwesenheiten:
            self.daten.abwesenheiten[tag] = set()
        if ma in self.daten.abwesenheiten[tag]:
            self.daten.abwesenheiten[tag].discard(ma)
        else:
            self.daten.abwesenheiten[tag].add(ma)
        self.refresh_tabelle()

    def _ganze_woche(self, ma, tage):
        alle_abw = all(ma in self.daten.abwesenheiten.get(t, set()) for t in tage)
        for t in tage:
            if t not in self.daten.abwesenheiten:
                self.daten.abwesenheiten[t] = set()
            if alle_abw:
                self.daten.abwesenheiten[t].discard(ma)
            else:
                self.daten.abwesenheiten[t].add(ma)
        self.refresh_tabelle()


# ═══════════════════════════════════════════════════════════════
#  Tab 4: Planung & Export
# ═══════════════════════════════════════════════════════════════

class PlanungTab(QWidget):
    def __init__(self, daten, abw_tab):
        super().__init__()
        self.daten = daten
        self.abw_tab = abw_tab
        self._letzter_plan = None
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)

        # Modus
        g_modus = QGroupBox("Planungsmodus")
        hl_m = QHBoxLayout(g_modus)
        self.rb_fair = QRadioButton("Fairer Algorithmus (automatische Verteilung)")
        self.rb_rot  = QRadioButton("Feste Rotation (eigene Reihenfolge)")
        self.rb_fair.setChecked(not self.daten.rotations_modus)
        self.rb_rot.setChecked(self.daten.rotations_modus)
        self.rb_fair.toggled.connect(lambda v: setattr(self.daten, 'rotations_modus', not v))
        hl_m.addWidget(self.rb_fair)
        hl_m.addWidget(self.rb_rot)
        hl_m.addStretch()
        btn_rot = sec_btn("Reihenfolge & Start festlegen")
        btn_rot.clicked.connect(self._rotation_editor)
        hl_m.addWidget(btn_rot)
        lay.addWidget(g_modus)

        # Aktions-Buttons
        g_act = QGroupBox("Aktionen")
        hl_a = QHBoxLayout(g_act)
        btn_plan = QPushButton("▶  Plan erstellen")
        btn_neu  = sec_btn("🔀  Neue Woche / neu würfeln")
        btn_farb = sec_btn("🎨  Taktfarben")
        btn_xl   = QPushButton("📊  Excel exportieren")
        btn_druck= sec_btn("🖨  Druckansicht")
        btn_plan.clicked.connect(self._plan_erstellen)
        btn_neu.clicked.connect(self._plan_neu)
        btn_farb.clicked.connect(self._farben_editor)
        btn_xl.clicked.connect(self._excel_exportieren)
        btn_druck.clicked.connect(self._druckansicht)
        for b in [btn_plan, btn_neu, btn_farb, btn_xl, btn_druck]:
            hl_a.addWidget(b)
        hl_a.addStretch()
        lay.addWidget(g_act)

        # Vorschau-Tabelle
        g_vor = QGroupBox("Vorschau")
        vl_v = QVBoxLayout(g_vor)
        self.tbl = QTableWidget()
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        self.tbl.setShowGrid(True)
        self.tbl.setGridStyle(Qt.PenStyle.SolidLine)
        vl_v.addWidget(self.tbl)
        lay.addWidget(g_vor, 1)

    def _aktive_tage(self):
        return self.abw_tab.aktive_tage()

    def _plan_erstellen(self):
        # Qualifikationen aus Tab 2 lesen
        tage = self._aktive_tage()
        if not tage:
            QMessageBox.warning(self, "Keine Tage", "Bitte mindestens einen Arbeitstag wählen.")
            return
        if not self.daten.alle_mas:
            QMessageBox.warning(self, "Keine Mitarbeiter", "Bitte Mitarbeiter anlegen.")
            return
        if not self.daten.takte:
            QMessageBox.warning(self, "Keine Takte", "Bitte Takte anlegen.")
            return

        # Warnung
        warnungen = []
        for takt in self.daten.takte:
            qual = [ma for ma in self.daten.alle_mas
                    if self.daten.ma_kann_takt(ma, takt) or not self.daten.qualifikationen.get(ma)]
            if len(qual) < self.daten.get_besetzung(takt):
                warnungen.append(f"'{takt}': {self.daten.get_besetzung(takt)}× benötigt, "
                                  f"nur {len(qual)} qualifiziert")
        if warnungen:
            ans = QMessageBox.question(self, "Warnung",
                "Mögliche Probleme:\n\n" + "\n".join(warnungen) + "\n\nTrotzdem erstellen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if ans != QMessageBox.StandardButton.Yes:
                return

        random.seed()
        if self.daten.rotations_modus:
            self._letzter_plan = plan_rotation(self.daten, tage)
        else:
            self._letzter_plan = plan_erstellen(self.daten, tage)
        # gesamt_zaehler wird NICHT hier aktualisiert – nur beim finalen Export/Archivieren
        self._vorschau_zeigen(tage)

    def _plan_neu(self):
        ans = QMessageBox.question(self, "Neue Woche",
            "Abwesenheiten zurücksetzen?\n\nJa = Alle auf 'Da'\nNein = Beibehalten",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ans == QMessageBox.StandardButton.Yes:
            self.daten.abwesenheiten = {}
            self.abw_tab.refresh_tabelle()
        self._plan_erstellen()

    def _vorschau_zeigen(self, tage):
        plan = self._letzter_plan
        mas = self.daten.alle_mas
        self.tbl.clear()
        self.tbl.setColumnCount(len(tage) + 1)
        self.tbl.setRowCount(len(mas))
        headers = ["Mitarbeiter"] + tage
        self.tbl.setHorizontalHeaderLabels(headers)

        # Header styling
        for ci in range(len(headers)):
            self.tbl.horizontalHeaderItem(ci).setForeground(QColor(P['text2']))

        for ri, ma in enumerate(mas):
            is_temp = self.daten.ist_temporaer(ma)
            # MA-Name Zelle
            name_item = QTableWidgetItem(f"{'▲ ' if is_temp else ''}{ma}")
            name_item.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            name_item.setForeground(QColor(P['gold'] if is_temp else P['text']))
            name_item.setBackground(QColor(P['bg4']))
            self.tbl.setItem(ri, 0, name_item)

            for ti, tag in enumerate(tage):
                eintrag = plan.get(tag, {}).get(ma)
                if eintrag == "ABWESEND":
                    itm = QTableWidgetItem("Abwesend")
                    itm.setBackground(QColor("#4A0A0A"))
                    itm.setForeground(QColor("#FF6B6B"))
                elif eintrag == "UEBERHANG":
                    itm = QTableWidgetItem("Überhang")
                    itm.setBackground(QColor("#2A2A1A"))
                    itm.setForeground(QColor("#C9A84C"))
                    itm.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
                elif not eintrag:
                    itm = QTableWidgetItem("—")
                    itm.setBackground(QColor(P['bg3']))
                    itm.setForeground(QColor(P['text2']))
                else:
                    itm = QTableWidgetItem(eintrag)
                    farbe = self.daten.takt_farbe(eintrag)
                    itm.setBackground(QColor(farbe))
                    itm.setForeground(QColor(P['white']))
                    itm.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
                itm.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tbl.setItem(ri, ti+1, itm)

        self.tbl.resizeRowsToContents()

    def _farben_editor(self):
        if not self.daten.takte:
            QMessageBox.warning(self, "Keine Takte", "Bitte zuerst Takte anlegen.")
            return
        from PyQt6.QtWidgets import QColorDialog
        dlg = QDialog(self)
        dlg.setWindowTitle("Taktfarben anpassen")
        dlg.setMinimumWidth(400)
        lay = QVBoxLayout(dlg)
        lay.addWidget(label("Klicke auf eine Farbe zum Ändern:", bold=True))
        btns = {}
        grid = QGridLayout()
        for i, takt in enumerate(self.daten.takte):
            lbl = QLabel(takt)
            lbl.setStyleSheet(f"color:{P['text']};padding:4px;")
            farbe = self.daten.takt_farbe(takt)
            btn = QPushButton()
            btn.setFixedSize(40, 28)
            btn.setStyleSheet(f"background:{farbe};border-radius:4px;border:1px solid {P['border']};")
            btn.clicked.connect(lambda _, t=takt, b=btn: self._pick_color(t, b))
            grid.addWidget(lbl, i, 0)
            grid.addWidget(btn, i, 1)
            btns[takt] = btn
        lay.addLayout(grid)
        btn_preset = sec_btn("Preset zurücksetzen")
        btn_preset.clicked.connect(lambda: self._preset_reset(btns))
        btn_close = QPushButton("Schließen")
        btn_close.clicked.connect(dlg.accept)
        hl = QHBoxLayout()
        hl.addWidget(btn_preset)
        hl.addWidget(btn_close)
        lay.addLayout(hl)
        dlg.exec()
        if self._letzter_plan:
            self._vorschau_zeigen(self._aktive_tage())

    def _pick_color(self, takt, btn):
        from PyQt6.QtWidgets import QColorDialog
        farbe = self.daten.takt_farbe(takt)
        col = QColorDialog.getColor(QColor(farbe), self, f"Farbe für '{takt}'")
        if col.isValid():
            self.daten.takt_farben[takt] = col.name()
            btn.setStyleSheet(f"background:{col.name()};border-radius:4px;border:1px solid {P['border']};")

    def _preset_reset(self, btns):
        for i, takt in enumerate(self.daten.takte):
            farbe = TAKT_PRESET[i % len(TAKT_PRESET)]
            self.daten.takt_farben[takt] = farbe
            btns[takt].setStyleSheet(f"background:{farbe};border-radius:4px;border:1px solid {P['border']};")

    def _excel_exportieren(self):
        if not self._letzter_plan:
            QMessageBox.warning(self, "Kein Plan", "Bitte zuerst einen Plan erstellen.")
            return
        heute = date.today()
        # Naechste Kalenderwoche
        naechste_woche = heute.isocalendar()[1] + 1
        jahr = heute.isocalendar()[0]
        if naechste_woche > 52:
            from datetime import timedelta
            letzter_tag = date(heute.year, 12, 28)
            max_kw = letzter_tag.isocalendar()[1]
            if naechste_woche > max_kw:
                naechste_woche = 1
                jahr += 1
        kw_str = f"KW{naechste_woche:02d}"
        datum = f"{jahr}_{kw_str}"
        pfad, _ = QFileDialog.getSaveFileName(
            self, "Excel speichern", os.path.join(_BASIS, f"Clicktakt_{kw_str}.xlsx"),
            "Excel (*.xlsx)")
        if pfad:
            tage = self._aktive_tage()
            self._excel_schreiben(pfad, tage)
            self.daten.gesamt_zaehler_aktualisieren(self._letzter_plan)
            self._archiv_auto(datum, tage)
            QMessageBox.information(self, "Gespeichert", f"Plan als {kw_str} gespeichert:\n{pfad}")

    def _excel_schreiben(self, pfad, tage):
        wb = Workbook()
        ws = wb.active
        ws.title = "Wochenplan"
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def z(row, col, wert, fett=False, bg=None, fg="FFFFFF"):
            c = ws.cell(row=row, column=col, value=wert)
            c.font = XFont(bold=fett, color=fg, name="Segoe UI", size=10)
            if bg:
                c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border

        ws.column_dimensions["A"].width = 18
        z(1, 1, "Mitarbeiter / Tag", fett=True, bg="111111")
        for ti, tag in enumerate(tage):
            z(1, ti+2, tag, fett=True, bg="1C1C1C")
            ws.column_dimensions[get_column_letter(ti+2)].width = 14
        ws.row_dimensions[1].height = 22

        for ri, ma in enumerate(self.daten.alle_mas):
            row = ri + 2
            ws.row_dimensions[row].height = 20
            z(row, 1, ma, fett=True, bg="2E2E2E")
            for ti, tag in enumerate(tage):
                eintrag = self._letzter_plan.get(tag, {}).get(ma)
                if eintrag == "ABWESEND":
                    z(row, ti+2, "Abwesend", bg="4A0A0A", fg="FF6B6B")
                elif eintrag == "UEBERHANG":
                    z(row, ti+2, "Überhang", bg="2A2A1A", fg="C9A84C")
                elif not eintrag:
                    z(row, ti+2, "—", bg="252525", fg="666666")
                else:
                    farbe = self.daten.takt_farbe(eintrag).lstrip("#")
                    z(row, ti+2, eintrag, bg=farbe)
        wb.save(pfad)

    def _archiv_auto(self, datum, tage):
        # datum ist hier bereits kw_str (z.B. "2026_KW21")
        if any(e.get("datum") == datum for e in self.daten.archiv):
            return
        plan_ser = {tag: dict(mas) for tag, mas in self._letzter_plan.items()}
        self.daten.archiv.insert(0, {
            "datum": datum, "tage": tage,
            "plan": plan_ser,
            "mitarbeiter": list(self.daten.alle_mas),
            "zaehler_snapshot": {
                ma: dict(z) for ma, z in self.daten.gesamt_zaehler.items()
            },
        })
        self.daten.archiv = self.daten.archiv[:8]

    def _druckansicht(self):
        if not self._letzter_plan:
            QMessageBox.warning(self, "Kein Plan", "Bitte zuerst einen Plan erstellen.")
            return
        import webbrowser, tempfile
        tage = self._aktive_tage()
        # Naechste Kalenderwoche berechnen
        heute = date.today()
        naechste_kw = heute.isocalendar()[1] + 1
        jahr = heute.isocalendar()[0]
        if naechste_kw > 52:
            from datetime import timedelta
            max_kw = date(heute.year, 12, 28).isocalendar()[1]
            if naechste_kw > max_kw:
                naechste_kw = 1
                jahr += 1
        kw_str = f"KW{naechste_kw:02d}"
        datum = kw_str
        zeilen = ""
        for ma in self.daten.alle_mas:
            temp_lbl = " <span style='color:#C9A84C'>(T)</span>" if self.daten.ist_temporaer(ma) else ""
            zellen = ""
            for tag in tage:
                e = self._letzter_plan.get(tag, {}).get(ma)
                if e == "ABWESEND":
                    zellen += "<td style='background:#4A0A0A;color:#FF6B6B'>Abwesend</td>"
                elif e == "UEBERHANG":
                    zellen += "<td style='background:#2A2A1A;color:#C9A84C;font-weight:bold'>Überhang</td>"
                elif not e:
                    zellen += "<td style='background:#252525;color:#666'>—</td>"
                else:
                    farbe = self.daten.takt_farbe(e)
                    zellen += f"<td style='background:{farbe};color:white;font-weight:bold'>{e}</td>"
            zeilen += f"<tr><td class='ma'>{ma}{temp_lbl}</td>{zellen}</tr>\n"
        th = "".join(f"<th>{t}</th>" for t in tage)
        html = (
            "<!DOCTYPE html><html><head><meta charset='UTF-8'>"
            f"<title>Clicktakt {datum}</title>"
            "<style>"
            "body{background:#111;color:#f0f0f0;font-family:'Segoe UI',Arial;margin:20px}"
            "h1{font-size:20px;color:#D5001C;letter-spacing:2px}"
            "p{color:#888;font-size:12px}"
            "table{border-collapse:collapse;width:100%;margin-top:16px}"
            "th,td{border:1px solid #333;padding:8px 12px;text-align:center;font-size:11px}"
            "th{background:#1C1C1C;color:#888;font-weight:bold}"
            ".ma{background:#2E2E2E;text-align:left;font-weight:bold;color:#F0F0F0;padding-left:12px}"
            "button{background:#D5001C;color:white;border:none;padding:8px 20px;border-radius:4px;"
            "cursor:pointer;font-size:13px;margin-bottom:16px}"
            "@media print{button{display:none}body{background:white;color:black}"
            "th{background:#eee;color:black}.ma{background:#f5f5f5;color:black}}"
            "</style></head><body>"
            f"<h1>Clicktakt &mdash; {kw_str}</h1>"
            f"<p>Werk Leipzig  ·  {kw_str}</p>"
            "<button onclick='window.print()'>Drucken</button>"
            f"<table><tr><th>Mitarbeiter</th>{th}</tr>{zeilen}</table>"
            "</body></html>"
        )
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8")
        tmp.write(html)
        tmp.close()
        webbrowser.open("file:///" + tmp.name.replace(os.sep, "/"))

    def _rotation_editor(self):
        if not self.daten.alle_mas or not self.daten.takte:
            QMessageBox.warning(self, "Fehler", "Bitte zuerst Mitarbeiter und Takte anlegen.")
            return
        if not self.daten.rotations_reihenfolge:
            self.daten.rotations_reihenfolge = list(self.daten.alle_mas)
        for ma in self.daten.alle_mas:
            if ma not in self.daten.rotations_reihenfolge:
                self.daten.rotations_reihenfolge.append(ma)
        self.daten.rotations_reihenfolge = [
            m for m in self.daten.rotations_reihenfolge if m in self.daten.alle_mas_set()]

        dlg = QDialog(self)
        dlg.setWindowTitle("Rotationsreihenfolge & Startposition")
        dlg.setMinimumSize(560, 500)
        outer = QVBoxLayout(dlg)
        outer.setSpacing(12)

        # Titel
        outer.addWidget(label("Rotationsreihenfolge festlegen", bold=True, size=12))
        outer.addWidget(divider())

        # Haupt-Layout: Links + Rechts
        lay = QHBoxLayout()
        lay.setSpacing(20)

        # ── Links: Reihenfolge ──
        left = QVBoxLayout()
        left.addWidget(label("Reihenfolge der MAs:", bold=True))
        left.addWidget(label("Bestimmt den Abstand zwischen MAs", dim=True))
        lb = QListWidget()
        lb.setMinimumHeight(250)
        reihe = list(self.daten.rotations_reihenfolge)
        lb.addItems(reihe)
        left.addWidget(lb)

        hl_mv = QHBoxLayout()
        btn_hoch = sec_btn("▲  Hoch")
        btn_run  = sec_btn("▼  Runter")

        def hoch():
            i = lb.currentRow()
            if i > 0:
                reihe[i-1], reihe[i] = reihe[i], reihe[i-1]
                lb.clear(); lb.addItems(reihe)
                lb.setCurrentRow(i-1)

        def runter():
            i = lb.currentRow()
            if i < len(reihe)-1:
                reihe[i+1], reihe[i] = reihe[i], reihe[i+1]
                lb.clear(); lb.addItems(reihe)
                lb.setCurrentRow(i+1)

        btn_hoch.clicked.connect(hoch)
        btn_run.clicked.connect(runter)
        hl_mv.addWidget(btn_hoch)
        hl_mv.addWidget(btn_run)
        left.addLayout(hl_mv)
        lay.addLayout(left)

        # ── Rechts: Startposition (optional) ──
        right = QVBoxLayout()
        right.addWidget(label("Startposition (optional):", bold=True))
        right.addWidget(label("Leer lassen = automatisch ab Takt 1", dim=True))
        right.addSpacing(4)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f"border: 1px solid {P['border']}; border-radius: 4px;")
        inner = QWidget()
        inner.setStyleSheet(f"background: {P['bg4']};")
        g = QGridLayout(inner)
        g.setContentsMargins(8, 8, 8, 8)
        g.setSpacing(6)

        # Header
        hdr_ma   = QLabel("Mitarbeiter")
        hdr_takt = QLabel("Starttakt")
        hdr_ma.setStyleSheet(f"color:{P['text2']};font-weight:bold;font-size:9pt;")
        hdr_takt.setStyleSheet(f"color:{P['text2']};font-weight:bold;font-size:9pt;")
        g.addWidget(hdr_ma,   0, 0)
        g.addWidget(hdr_takt, 0, 1)

        start_combos = {}
        for i, ma in enumerate(self.daten.alle_mas):
            ma_lbl = QLabel(ma)
            ma_lbl.setStyleSheet(f"color:{P['text']};padding: 2px 4px;")
            cb = QComboBox()
            # Ersten Eintrag als "Automatisch" belassen
            cb.addItem("— Automatisch —")
            cb.addItems(self.daten.takte)
            # Gespeicherten Wert wiederherstellen
            if ma in self.daten.rotations_start and self.daten.rotations_start[ma] in self.daten.takte:
                cb.setCurrentText(self.daten.rotations_start[ma])
            else:
                cb.setCurrentIndex(0)  # Automatisch
            g.addWidget(ma_lbl, i+1, 0)
            g.addWidget(cb,     i+1, 1)
            start_combos[ma] = cb

        scroll.setWidget(inner)
        right.addWidget(scroll)
        lay.addLayout(right)
        outer.addLayout(lay)

        outer.addWidget(divider())

        # ── Bestätigungs-Buttons ──
        btn_row = QHBoxLayout()
        btn_abbrechen = sec_btn("Abbrechen")
        btn_ok = QPushButton("✓  Reihenfolge speichern")

        def ok():
            self.daten.rotations_reihenfolge = list(reihe)
            # Nur explizit gesetzte Starttakte speichern
            self.daten.rotations_start = {}
            for ma, cb in start_combos.items():
                if cb.currentIndex() > 0:  # Index 0 = "Automatisch"
                    self.daten.rotations_start[ma] = cb.currentText()
            QMessageBox.information(dlg, "Gespeichert",
                "Reihenfolge und Startposition wurden gespeichert.")
            dlg.accept()

        btn_abbrechen.clicked.connect(dlg.reject)
        btn_ok.clicked.connect(ok)
        btn_row.addWidget(btn_abbrechen)
        btn_row.addStretch()
        btn_row.addWidget(btn_ok)
        outer.addLayout(btn_row)

        dlg.exec()


# ═══════════════════════════════════════════════════════════════
#  Tab 5: Wochenverlauf
# ═══════════════════════════════════════════════════════════════

class ArchivTab(QWidget):
    def __init__(self, daten, planung_tab=None):
        super().__init__()
        self.daten = daten
        self.planung_tab = planung_tab  # Referenz auf PlanungTab fuer letzter_plan
        self._build()

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)

        top = QHBoxLayout()
        top.addWidget(label("Gespeicherte Wochenpläne (max. 8)", bold=True))
        top.addStretch()
        btn_ges = sec_btn("Gesamtverteilung anzeigen")
        btn_ges.clicked.connect(self._gesamtverteilung)
        btn_arc = QPushButton("Aktuellen Plan archivieren")
        btn_arc.clicked.connect(self._manuell_archivieren)
        btn_del = danger_btn("Ausgewählten löschen")
        btn_del.clicked.connect(self._loeschen)
        top.addWidget(btn_ges)
        top.addWidget(btn_arc)
        top.addWidget(btn_del)
        lay.addLayout(top)

        splitter = QSplitter(Qt.Orientation.Vertical)

        # Archiv-Liste
        self.lb = QListWidget()
        self.lb.setMaximumHeight(160)
        self.lb.currentRowChanged.connect(self._anzeigen)
        self.lb.keyPressEvent = lambda e: (self._loeschen() if e.key() == Qt.Key.Key_Delete else QListWidget.keyPressEvent(self.lb, e))
        splitter.addWidget(self.lb)

        # Vorschau
        g = QGroupBox("Vorschau")
        vl = QVBoxLayout(g)
        self.tbl = QTableWidget()
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tbl.verticalHeader().setVisible(False)
        vl.addWidget(self.tbl)
        splitter.addWidget(g)
        lay.addWidget(splitter, 1)
        self.refresh()

    def set_planung_tab(self, tab):
        self.planung_tab = tab

    def _manuell_archivieren(self):
        if not self.planung_tab or not self.planung_tab._letzter_plan:
            QMessageBox.warning(self, "Kein Plan",
                "Bitte zuerst einen Plan in Tab 'Planung & Export' erstellen.")
            return
        heute = date.today()
        naechste_kw = heute.isocalendar()[1] + 1
        jahr = heute.isocalendar()[0]
        if naechste_kw > 52:
            max_kw = date(heute.year, 12, 28).isocalendar()[1]
            if naechste_kw > max_kw:
                naechste_kw = 1
                jahr += 1
        kw_str = f"{jahr}_KW{naechste_kw:02d}"
        tage = self.planung_tab._aktive_tage()
        plan_ser = {tag: dict(mas) for tag, mas in self.planung_tab._letzter_plan.items()}
        if any(e.get("datum") == kw_str for e in self.daten.archiv):
            if QMessageBox.question(self, "Bereits vorhanden",
                f"{kw_str} ist bereits im Archiv. Ersetzen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) != QMessageBox.StandardButton.Yes:
                return
            self.daten.archiv = [e for e in self.daten.archiv if e.get("datum") != kw_str]
        eintrag = {
            "datum": kw_str,
            "tage": tage,
            "plan": plan_ser,
            "mitarbeiter": list(self.daten.alle_mas),
            "zaehler_snapshot": {
                ma: dict(z) for ma, z in self.daten.gesamt_zaehler.items()
            },
        }
        self.daten.archiv.insert(0, eintrag)
        self.daten.archiv = self.daten.archiv[:8]
        # Gesamtzaehler nur einmal pro Woche aktualisieren
        if self.planung_tab and self.planung_tab._letzter_plan:
            self.daten.gesamt_zaehler_aktualisieren(self.planung_tab._letzter_plan)
        self.refresh()
        QMessageBox.information(self, "Archiviert", f"Plan für {kw_str} wurde archiviert.")

    def refresh(self):
        self.lb.clear()
        for e in self.daten.archiv:
            self.lb.addItem(f"  {e['datum']}   ({', '.join(e['tage'])})")

    def _loeschen(self):
        i = self.lb.currentRow()
        if i >= 0:
            if QMessageBox.question(self, "Löschen", "Diesen Eintrag löschen?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            ) == QMessageBox.StandardButton.Yes:
                del self.daten.archiv[i]
                # Wenn noch Eintraege vorhanden: gesamt_zaehler auf letzten Snapshot
                if self.daten.archiv and "zaehler_snapshot" in self.daten.archiv[0]:
                    snap = self.daten.archiv[0]["zaehler_snapshot"]
                    self.daten.gesamt_zaehler = {ma: dict(z) for ma, z in snap.items()}
                self.refresh()
                self.tbl.clear()

    def _anzeigen(self, i):
        if i < 0 or i >= len(self.daten.archiv): return
        e = self.daten.archiv[i]
        plan, tage, mas = e["plan"], e["tage"], e["mitarbeiter"]
        self.tbl.clear()
        self.tbl.setColumnCount(len(tage)+1)
        self.tbl.setRowCount(len(mas))
        self.tbl.setHorizontalHeaderLabels(["Mitarbeiter"] + tage)
        alle_takte = list(dict.fromkeys(
            v for td in plan.values() for v in td.values() if v and v != "ABWESEND"))
        for ri, ma in enumerate(mas):
            ni = QTableWidgetItem(ma)
            ni.setBackground(QColor(P['bg4']))
            ni.setForeground(QColor(P['text']))
            ni.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            self.tbl.setItem(ri, 0, ni)
            for ti, tag in enumerate(tage):
                val = plan.get(tag, {}).get(ma)
                if val == "ABWESEND":
                    itm = QTableWidgetItem("Abwesend")
                    itm.setBackground(QColor("#4A0A0A"))
                    itm.setForeground(QColor("#FF6B6B"))
                elif val == "UEBERHANG":
                    itm = QTableWidgetItem("Überhang")
                    itm.setBackground(QColor("#2A2A1A"))
                    itm.setForeground(QColor("#C9A84C"))
                    itm.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
                elif not val:
                    itm = QTableWidgetItem("—")
                    itm.setBackground(QColor(P['bg3']))
                    itm.setForeground(QColor(P['text2']))
                else:
                    itm = QTableWidgetItem(val)
                    farbe = self.daten.takt_farbe(val)
                    itm.setBackground(QColor(farbe))
                    itm.setForeground(QColor(P['white']))
                    itm.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
                itm.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tbl.setItem(ri, ti+1, itm)
        self.tbl.resizeRowsToContents()

    def _gesamtverteilung(self):
        if not self.daten.gesamt_zaehler:
            QMessageBox.information(self, "Keine Daten",
                "Noch keine Daten. Bitte erst Pläne erstellen.")
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("Gesamtverteilung aller Wochen")
        dlg.resize(700, 400)
        lay = QVBoxLayout(dlg)
        tbl = QTableWidget()
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setColumnCount(len(self.daten.takte)+2)
        tbl.setRowCount(len(self.daten.alle_mas))
        tbl.setHorizontalHeaderLabels(["Mitarbeiter"] + self.daten.takte + ["Gesamt"])
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        tbl.verticalHeader().setVisible(False)
        for ri, ma in enumerate(self.daten.alle_mas):
            ni = QTableWidgetItem(ma)
            ni.setBackground(QColor(P['bg4']))
            ni.setForeground(QColor(P['text']))
            tbl.setItem(ri, 0, ni)
            gz = self.daten.gesamt_zaehler.get(ma, {})
            werte = [gz.get(t, 0) for t in self.daten.takte]
            gesamt = sum(werte)
            mx = max(werte) if werte else 0
            mn = min(werte) if werte else 0
            for ti, n in enumerate(werte):
                itm = QTableWidgetItem(str(n))
                itm.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if mx > 0 and n == mx:
                    itm.setBackground(QColor("#4A1A1A"))
                    itm.setForeground(QColor("#FF6B6B"))
                elif mx > 0 and n == mn:
                    itm.setBackground(QColor("#0A3D1A"))
                    itm.setForeground(QColor("#4CAF50"))
                else:
                    itm.setBackground(QColor(P['bg3']))
                    itm.setForeground(QColor(P['text']))
                tbl.setItem(ri, ti+1, itm)
            gi = QTableWidgetItem(str(gesamt))
            gi.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            gi.setBackground(QColor(P['bg4']))
            gi.setForeground(QColor(P['text2']))
            tbl.setItem(ri, len(self.daten.takte)+1, gi)
        lay.addWidget(tbl)
        btn = QPushButton("Schließen")
        btn.clicked.connect(dlg.accept)
        lay.addWidget(btn)
        dlg.exec()


# ═══════════════════════════════════════════════════════════════
#  Haupt-Fenster
# ═══════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Clicktakt  |  Werk Leipzig")
        self.resize(1200, 780)
        self.setMinimumSize(900, 600)

        self.daten = TeamDaten()
        if os.path.exists(SAVE_FILE):
            try:
                self.daten.laden()
                # Abwesenheiten pruefen: wenn letzte KW veraltet ist -> zuruecksetzen
                self._abwesenheiten_pruefen()
            except Exception as e:
                backup = SAVE_FILE + ".backup"
                hat_backup = os.path.exists(backup)
                msg = (f"Konfiguration konnte nicht geladen werden:\n{e}\n\n"
                       + ("Ein Backup wurde gefunden. Backup wiederherstellen?"
                          if hat_backup else
                          "Das Programm startet mit leeren Daten."))
                if hat_backup:
                    ans = QMessageBox.question(None, "Ladefehler", msg,
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
                    if ans == QMessageBox.StandardButton.Yes:
                        try:
                            self.daten.backup_wiederherstellen()
                        except Exception as e2:
                            QMessageBox.warning(None, "Backup fehlgeschlagen",
                                f"Backup konnte nicht geladen werden:\n{e2}")
                else:
                    QMessageBox.warning(None, "Ladefehler", msg)

        central = QWidget()
        self.setCentralWidget(central)
        main_lay = QVBoxLayout(central)
        main_lay.setContentsMargins(0, 0, 0, 0)
        main_lay.setSpacing(0)

        # Header
        main_lay.addWidget(HeaderWidget())

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        main_lay.addWidget(self.tabs)

        self.tab_stamm = StammdatenTab(self.daten)
        self.tab_quali  = QualifikationenTab(self.daten)
        self.tab_abw    = AbwesenheitenTab(self.daten)
        self.tab_plan   = PlanungTab(self.daten, self.tab_abw)
        self.tab_archiv = ArchivTab(self.daten)

        self.tabs.addTab(self.tab_stamm,  "  Stammdaten  ")
        self.tabs.addTab(self.tab_quali,  "  Qualifikationen  ")
        self.tabs.addTab(self.tab_abw,    "  Abwesenheiten  ")
        self.tabs.addTab(self.tab_plan,   "  Planung && Export  ")
        self.tabs.addTab(self.tab_archiv, "  Wochenverlauf  ")

        self.tab_archiv.set_planung_tab(self.tab_plan)
        self.tab_stamm.changed.connect(self._on_changed)
        # Abwesenheitstabelle nach Laden aktualisieren
        self.tab_abw.refresh_tabelle()

    def _on_changed(self):
        self.tab_quali.refresh()
        self.tab_abw.refresh_tabelle()

    def closeEvent(self, e):
        self.tab_quali.speichern()
        self.daten.speichern()
        e.accept()


# ═══════════════════════════════════════════════════════════════
#  Start
# ═══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(QSS)
    # Dark Palette
    palette = QPalette()
    palette.setColor(QPalette.ColorRole.Window,       QColor(P['bg']))
    palette.setColor(QPalette.ColorRole.WindowText,   QColor(P['text']))
    palette.setColor(QPalette.ColorRole.Base,         QColor(P['bg4']))
    palette.setColor(QPalette.ColorRole.Text,         QColor(P['text']))
    palette.setColor(QPalette.ColorRole.Button,       QColor(P['bg4']))
    palette.setColor(QPalette.ColorRole.ButtonText,   QColor(P['text']))
    palette.setColor(QPalette.ColorRole.Highlight,    QColor(P['red']))
    palette.setColor(QPalette.ColorRole.HighlightedText, QColor(P['white']))
    app.setPalette(palette)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())
